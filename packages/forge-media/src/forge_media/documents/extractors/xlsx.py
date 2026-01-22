"""XLSX spreadsheet extraction using openpyxl."""

from __future__ import annotations

from io import BytesIO

from forge_media.documents.extractors.base import (
    DocumentExtractor,
    DocumentPage,
    ExtractedDocument,
    ExtractedTable,
    ExtractionError,
    ExtractionOptions,
)


class XLSXExtractor(DocumentExtractor):
    """
    Extract content from Microsoft Excel (.xlsx) spreadsheets.

    Each sheet is treated as a separate "page" with its data represented
    as a table.

    Example:
        ```python
        extractor = XLSXExtractor()
        with open("spreadsheet.xlsx", "rb") as f:
            data = f.read()

        result = await extractor.extract(data)
        for page in result.pages:
            print(f"Sheet: {page.metadata.get('sheet_name')}")
            for table in page.tables:
                print(f"  Rows: {table.row_count}")
        ```
    """

    @property
    def supported_mime_types(self) -> list[str]:
        return [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",  # Legacy .xls (limited support)
        ]

    async def extract(
        self,
        data: bytes,
        options: ExtractionOptions | None = None,
    ) -> ExtractedDocument:
        """
        Extract content from an XLSX spreadsheet.

        Args:
            data: Raw XLSX bytes
            options: Extraction options

        Returns:
            ExtractedDocument with each sheet as a page
        """
        options = options or ExtractionOptions()

        try:
            import openpyxl
        except ImportError as e:
            raise ExtractionError(
                "openpyxl is required for XLSX extraction. "
                "Install with: pip install openpyxl",
                cause=e,
            )

        try:
            workbook = openpyxl.load_workbook(BytesIO(data), data_only=True)
        except Exception as e:
            raise ExtractionError(f"Failed to read XLSX: {e}", cause=e)

        pages: list[DocumentPage] = []
        tables: list[ExtractedTable] = []
        text_parts: list[str] = []

        for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
            # Respect max_pages limit
            if options.max_pages and sheet_idx >= options.max_pages:
                break

            sheet = workbook[sheet_name]

            # Extract rows from sheet
            rows: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                # Skip completely empty rows
                if any(cell.strip() for cell in row_data):
                    rows.append(row_data)

            if not rows:
                continue

            # First row as headers, rest as data
            headers = rows[0]
            data_rows = rows[1:]

            # Create table for this sheet
            table = ExtractedTable(
                page_number=sheet_idx + 1,
                table_index=len(tables),
                headers=headers,
                rows=data_rows,
            )
            tables.append(table)

            # Create text representation of the sheet
            sheet_text = f"Sheet: {sheet_name}\n"
            sheet_text += "\t".join(headers) + "\n"
            for row in data_rows:
                sheet_text += "\t".join(row) + "\n"
            text_parts.append(sheet_text)

            # Create page for this sheet
            pages.append(
                DocumentPage(
                    page_number=sheet_idx + 1,
                    text=sheet_text,
                    tables=[table],
                    images=[],
                    metadata={"sheet_name": sheet_name},
                )
            )

        full_text = "\n\n".join(text_parts)

        # Extract workbook metadata
        metadata = {}
        if options.extract_metadata:
            props = workbook.properties
            if props:
                metadata = {
                    "title": props.title,
                    "creator": props.creator,
                    "subject": props.subject,
                    "description": props.description,
                    "created": props.created.isoformat() if props.created else None,
                    "modified": props.modified.isoformat() if props.modified else None,
                    "sheet_names": workbook.sheetnames,
                }
                # Remove None values
                metadata = {k: v for k, v in metadata.items() if v is not None}

        return ExtractedDocument(
            pages=pages,
            full_text=full_text,
            tables=tables,
            metadata=metadata,
            page_count=len(pages),
            word_count=len(full_text.split()),
        )
